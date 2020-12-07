#!/usr/bin/env python3
#----------------------------------------------------------
#   Name: style.py
#   Author: xyy15926
#   Created at: 2020-11-12 20:55:46
#   Updated at: 2020-11-12 20:58:20
#   Description: 
#     Set style for excel output of dataframe:
#       1. fit only when using `openpyxl` as engine
#----------------------------------------------------------

# %%
from openpyxl.styles import (Font, Alignment, Border, Side,
            PatternFill, NamedStyle)
import openpyxl as pyxl
from typing import (Any, Tuple, Union)
from ..drawer.color import (LIGHT_COLORS, DEEP_COLORS)
import re

# %%

# Define some customed style here

# FONTS
FONT_EMPHASIZED = Font(
    name="微软雅黑",
    size=14,
    bold=True,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color="00000000"
)
FONT_EMPHASIZED_R = Font(
    name="微软雅黑",
    size=14,
    bold=True,
    color="FFFFFF"
)
FONT_CONTENT = Font(
    name="微软雅黑",
    size=10,
    color="00000000"
)
FONT_CONTENT_R = Font(
    name="微软雅黑",
    size=10,
    color="FFFFFF"
)
FONT_HYPERLINK = Font(
    name="微软雅黑",
    size=10,
    italic=True,
    underline="singleAccounting",
    color="2980b9"
)

# FILL
FILL_DARK = PatternFill(
    fill_type="solid",
    start_color="2c3e50"
)
FILL_GREY = PatternFill(
    fill_type="solid",
    start_color="d5d8dc"
)
FILL_COLD = PatternFill(
    fill_type="solid",
    start_color="1a5265"
)
FILL_WARM = PatternFill(
    fill_type="solid",
    start_color="a04000"
)

# SIDE
SIDE_THICK = Side(
    border_style="thick",
    color="000000"
)
SIDE_THIN = Side(
    style="thin",
    color="000000"
)

# BORDER
BORDER_THICK = Border(
    left=SIDE_THICK,
    right=SIDE_THICK,
    top=SIDE_THICK,
    bottom=SIDE_THICK
)
BORDER_THIN = Border(
    left=SIDE_THIN,
    right=SIDE_THIN,
    top=SIDE_THIN,
    bottom=SIDE_THIN
)

# ALIGNMENT
ALIGN_WRAP = Alignment(
    horizontal="center",
    vertical="center",
    text_rotation=0,
    wrap_text=True,
    shrink_to_fit=False,
    indent=0
)
ALIGN_NOWRAP = Alignment(
    horizontal="center",
    vertical="center",
    text_rotation=0,
    wrap_text=False,
    shrink_to_fit=True,
    indent=0
)
ALIGN_LEFT = Alignment(
    horizontal="left",
    vertical="center",
    wrap_text=True,
    shrink_to_fit=True
)

# %%
# define named-styles' names
HEADER_STYLE = "header"
CONTENT_STYLE = "content"
CONTENT_NOFILL = "content_nofill"
CONTENT_LEFT = "content_left"

# define named-styles
STYLES = {
    HEADER_STYLE: {
        "font": FONT_EMPHASIZED_R,
        "border": BORDER_THICK,
        "fill": FILL_DARK,
        "alignment": ALIGN_NOWRAP
    },
    CONTENT_STYLE: {
        "font": FONT_CONTENT,
        "fill": FILL_GREY,
        "border": BORDER_THIN,
        "alignment": ALIGN_WRAP
    },
    CONTENT_NOFILL: {
        "font": FONT_CONTENT,
        "border": BORDER_THIN,
        "alignment": ALIGN_WRAP
    },
    CONTENT_LEFT: {
        "font": FONT_CONTENT,
        "border": BORDER_THIN,
        "alignment": ALIGN_LEFT
    }
}


# %%
def add_format(
    workbook: Any,
) -> None:
    """
    Description:
    1. Add predef styles in this function to workbook

    Params:
    workbook: openpyxl workbook

    Return:
    None
    """
    # add named-styles to `workbook`
    for sty_name, stys in STYLES.items():
        named_sty = NamedStyle(sty_name)
        for attr_name, attr_val in stys.items():
            setattr(named_sty, attr_name, attr_val)
        workbook.add_named_style(named_sty)

#%%
def _get_col_num(
    code: str
) -> int:
    """
    Description:
    Get column's number index acording to its code index

    Params:
    code: "A","AB", etc
    
    return:
    number index
    """
    num = 0
    while len(code) > 0:
        num = num * 26 + ord(code[0]) - ord("A") + 1
        code = code[1:]
    return num

def _get_col_str(
    num: int
) -> str:
    """
    Description:
    Get column's code index acording to its number index

    Params:
    number index
    
    return:
    code: "A","AB", etc
    """
    str_ = ""
    while num > 0:
        a, b = divmod(num-1, 26)
        str_ = chr(ord("A") + b) + str_
        num = a
    return str_

def _split_cell(
    cell_code: str
) -> Tuple[str, int]:
    """
    Decription:
    Split cell's index into columns index and row index

    Params:
    cell_code: "A1", "BB12", etc

    Return:
    (column index, row index)
    """
    rets = re.match("([A-Z]+)(\d+)", cell_code).groups()
    return rets[0], int(rets[1])

def _con_cell(
    cell_idx: Union[tuple, list]
) -> str:
    """
    Description:
    Construct cell code acording to a list with column index
        row index

    Params:
    cell_idx: [column_idx, row_idx]

    Return:
    cell code
    """
    if isinstance(cell_idx[0], int):
        return f"{_get_col_str(cell_idx[0])}{cell_idx[1]}"
    else:
        return f"{cell_idx[0]}{cell_idx[1]}"

def _around_cell(
    cell_code: str,
    move: [0, 0]
) -> str:
    """
    Description:
    Get the index of cell relative to `cell` specified by `cell_code`

    Params:
    move: [column_movement, row_movement]
        positive: move right or down
        negtive: move left or up

    Return:
    cell code
    """
    c, r = _split_cell(cell_code)
    c = _get_col_str(_get_col_num(c) + move[0])
    r += move[1]
    return f"{c}{r}"

# %%
def apply_style(
    ws: pyxl.worksheet.worksheet.Worksheet,
    start: str,
    end: str,
    named_style: Union[str, dict],
    keep:bool = True
) -> None:
    """
    Description:
    Apply style according to `name_style` to `ws[start: end]`

    Params:
    ws: worksheet
    start: start cell of the range
    end: end cell of the range
    name_style: if `str` is passed, `SYTLES` will be look up for it
    keep: if to keep the original style that not defined in the cells

    Return:
    None
    """
    # apply named-style directly
    if not keep:
        for cells in ws[start: end]:
            if isinstance(cells, tuple):
                for cell in cells:
                    cell.style = named_style
            else:
                cells.style = named_style
    # set each style elements in names-styles for cells
    else:
        if isinstance(named_style, str):
            named_style = STYLES[named_style]
        for cells in ws[start: end]:
            if isinstance(cells, tuple):
                for cell in cells:
                    for attr_name, attr_val in named_style.items():
                        setattr(cell, attr_name, attr_val)
            else:
                for attr_name, attr_val in named_style.items():
                    setattr(cell, attr_name, attr_val)

# %%
def format_worksheet(
    ws: pyxl.worksheet.worksheet.Worksheet,
    anchors:list = ["B2", "C3", ""],
    keep=True,
    set_header:bool = True,
    set_content:bool = True,
    set_width:bool = True,
    frozen:Union[bool, str, tuple, list] = True
) -> None:
    """
    Description:

    Params:
    ws: worksheet to be formatted
    anchors: [CR1, CR2, CRN], which determine the whole df
        CR1                         col_header
                    CR2



        row_header                  CRN

    Return:
    None
    """
    # Set sheet' style
    ws.sheet_view.showGridLines=False
    ws.sheet_view.showRowColHeaders=True

    # Copy to avoid affecting the list
    anchors = anchors.copy()

    # Set data dimensions with default dimensions
    if not anchors:
        cr1, crn = ws.dimensions.split(":")
        anchors = [cr1, cr1, crn]
    if anchors[2] == "":
        anchors[2] = ws.dimensions.split(":")[1]

    # Get data structure in `ws`
    # ach:[
    #   [cr1_c, cr1_r],
    #   [cr2_c, cr2_r],
    #   [crn_c, crn_r],
    # ]
    ach = [0] * 3
    for idx, anchor in enumerate(anchors):
        if isinstance(anchor, str):
            cr_c, cr_r = _split_cell(anchor)
        else:
            cr_c, cr_r = anchor, _get_col_str(anchor[1])
        ach[idx] = [cr_c, cr_r]

    # Row header exists
    if set_header:
        if ach[0][0] != ach[1][0]:
            apply_style(
                ws,
                _con_cell(ach[0]),
                _around_cell(_con_cell([ach[1][0], ach[2][1]]), (-1, 0)),
                HEADER_STYLE,
                keep
            )
        # column header exists
        if ach[0][1] != ach[1][1]:
            apply_style(
                ws,
                _con_cell(ach[0]),
                _around_cell(_con_cell([ach[2][0], ach[1][1]]), (0, -1)),
                HEADER_STYLE,
                keep
            )
    # Set content style
    if set_content:
        apply_style(
            ws,
            _con_cell(ach[1]),
            _con_cell(ach[2]),
            CONTENT_NOFILL,
            keep
        )

    # Set freeze panes
    if frozen:
        if frozen == True:
            frozen = anchors[1]
        elif isinstance(frozen, list) or isinstance(frozen, tuple):
            frozen = _con_cell(frozen)
        ws.freeze_panes = frozen

    # Set columns' width, between 10-100, acording to value length
    # It' seems that `width = len * font_size / 10` works fine, for
        # ASCII chars
    # Cell has valid `font` attribute with no-None value. But if cell
        # was set with any other styles beyond `openpyxl`, the `font`
        # attribute's value will be set to `None`
    if set_width:
        start_column_idx = _get_col_num(ach[0][0])
        end_column_idx = _get_col_num(ach[2][0])
        for column_idx, column in enumerate(ws.iter_cols(start_column_idx, end_column_idx), start_column_idx):
            widths = []
            for cell in column:
                width = len(str(cell.value)) * cell.font.size // 20
                widths.append(width)
            width = min(max(*widths, 10), 100)
            ws.column_dimensions[_get_col_str(column_idx)].width = width + 2


# %%
def format_excel(
    wb: pyxl.workbook.workbook.Workbook,
    add_index:Union[bool, None] = None,
    formats: Union[bool, list] = True,
    anchors:list = ["B2", "C3", ""],
) -> None:
    """
    Description:
    Format excel workbook
    1. add index sheet with hyperlink, if specified
    2. format sheets in workbook

    Params:
    wb: workbook
    add_index: whether to add index sheet
        true/false: add/don't add index sheet
        None: add index sheet if 5 or more sheets in `wb`
    formats:
        true/false: whether format sheets
        list: sheets-name to be formated
    anchors: list[achors]/dict/list[str] specified the structure
            of the data in the sheets
        list[str]: the same `anchors` will be applyed for all sheets
        list[anchors]: `_anchors` in `anchors` will be applyed for
            sheets in the order of sheets in workbook
        dict[key, anchors]: `anchors` will applyed for sheets will
            the name just the `key`

    Return:
    None
    """
    # add index-page if `index_page` is set or more that 4 sheets
    # in the workbook, then
    # 1. add index item, excluding first index page
    # 2. format
    INDEX_CONST = "目录"
    if add_index or (add_index is None and len(wb.sheetnames) > 4):
        index_sheet = wb.create_sheet(INDEX_CONST, 0)
        index_sheet["B2"].value = INDEX_CONST
        index_sheet["C2"].value = "Description"
        for enum_idx, sheet_name in enumerate(wb.sheetnames[1:], 3):
            if sheet_name != INDEX_CONST:
                index_sheet[f"B{enum_idx}"].hyperlink = f"#{sheet_name}!B2"
                index_sheet[f"B{enum_idx}"].value = f"{sheet_name}"
        format_worksheet(index_sheet, ["B2", "B3", ""])

    # travese all worksheets, for each of them
    # 1. format
    # 2. add hyper-link to back to index page
    if formats is False:
        formats = []
    elif formats is True:
        formats = wb.sheetnames
        if INDEX_CONST in formats:
            formats.remove(INDEX_CONST)

    def add_back(ws):
        if ws.title != INDEX_CONST:
            if ws["A1"].value is None:
                ws["A1"].value = f'=HYPERLINK("!{INDEX_CONST}#B2", "BACK")'
                ws["A1"].font = FONT_HYPERLINK
            else:
                start_cell, end_cell = ws.dimensions.split(":")
                end_col, end_row = _split_cell(end_cell)
                ws[f"A{end_row}"].value = f'=HYPERLINK("!{INDEX_CONST}#B2", "BACK")'
                ws[f"A{end_row}"].FONT_HYPERLINK

    # format worksheet according to anchors
    if isinstance(anchors, dict):
        for ws_name in formats:
            if ws_name not in anchors:
                format_worksheet(wb[ws_name])
            else:
                format_worksheet(wb[ws_name], anchors[ws_name])
            if INDEX_CONST in wb.sheetnames:
                add_back(wb[ws_name])
    else:
        if isinstance(anchors[0], str):
            anchors = [anchors] * len(formats)
        for _anchors, ws_name in zip(anchors, formats):
            format_worksheet(wb[ws_name], _anchors)
            if INDEX_CONST in wb.sheetnames:
                add_back(wb[ws_name])


# %%
